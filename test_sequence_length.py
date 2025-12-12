#!/usr/bin/env python3
# 测试序列长度显示功能

import os
import sys
import tempfile
import pandas as pd
from openpyxl import Workbook

# 添加当前目录到路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import parser
import xml_generator

# 测试序列：包含修饰符的序列
test_sequence = "CmsUfsAmCmUmCfNCmAmGmCmAmGmAfCmAfCfUmGmGmGmsAmsUf"
print(f"测试序列: {test_sequence}")
print(f"原始序列长度: {len(test_sequence)}")

# 创建临时Excel文件用于测试
def create_test_excel():
    wb = Workbook()
    
    # 创建basicdata工作表
    ws_basic = wb.create_sheet("basicdata")
    ws_basic.append(["field", "value"])
    ws_basic.append(["ApplicantFileReference", "test_sequence_length"])
    ws_basic.append(["ApplicantName", "测试用户"])
    ws_basic.append(["ApplicantNameLatin", "Test User"])
    ws_basic.append(["InventorName", "测试发明人"])
    ws_basic.append(["InventorNameLatin", "Test Inventor"])
    ws_basic.append(["InventionTitle", "测试发明"])
    ws_basic.append(["earliestpriorityIPOfficeCode", "CN"])
    ws_basic.append(["ApplicationNumberText", "CN20240000000"])
    ws_basic.append(["earliestpriorityFilingDate", "2024-01-01"])
    
    # 创建seqdata工作表
    ws_seq = wb.create_sheet("seqdata")
    ws_seq.append(["序列", "分子类型", "来源", "修饰类型", "freetext1"])
    ws_seq.append([test_sequence, "RNA", "synthetic construct", "other RNA", "unknown nucleotide"])
    
    # 删除默认工作表
    wb.remove(wb["Sheet"])
    
    # 保存临时文件
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

# 运行测试
temp_excel = create_test_excel()
try:
    # 解析序列
    sequences = parser.read_sequences_from_excel(temp_excel)
    basic_data = parser.read_basic_data_from_excel(temp_excel)
    
    # 获取序列摘要
    sequence_summary = parser.get_sequence_summary(sequences)
    print(f"\n序列摘要信息:")
    print(f"总序列数: {sequence_summary['total_count']}")
    print(f"序列详细信息: {sequence_summary['details']}")
    
    # 生成XML
    print("\n生成XML...")
    xml_root = xml_generator.generate_xml(sequences, basic_data, ".")
    
    # 检查XML中的序列长度
    print("\n检查XML中的序列长度:")
    for seq_data in xml_root.findall("./SequenceData/INSDSeq"):
        seq_length = seq_data.find("INSDSeq_length").text
        seq_sequence = seq_data.find("INSDSeq_sequence").text
        source_location = seq_data.find("./INSDSeq_feature-table/INSDFeature[INSDFeature_key='source']/INSDFeature_location").text
        
        print(f"INSDSeq_length: {seq_length}")
        print(f"序列内容: {seq_sequence}")
        print(f"source特征位置: {source_location}")
        
        # 验证长度是否正确
        if int(seq_length) == len(test_sequence):
            print("✅ 序列长度正确: XML中显示的长度与原始序列长度一致")
        else:
            print(f"❌ 序列长度错误: XML中显示的长度({seq_length})与原始序列长度({len(test_sequence)})不一致")
            
        if source_location.endswith(str(len(test_sequence))):
            print("✅ source特征位置正确")
        else:
            print(f"❌ source特征位置错误")
            
finally:
    # 清理临时文件
    os.unlink(temp_excel)